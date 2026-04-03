import io
from datetime import datetime
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

class ExportManager:
    @staticmethod
    def to_excel(data, period_name="Rapport"):
        """Génère un fichier Excel stylisé."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Résumé Financier"

        # Styles
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # En-tête du document
        ws.merge_cells('A1:D1')
        ws['A1'] = f"MKARIBU - BILAN COMPTABLE ({period_name.upper()})"
        ws['A1'].font = Font(size=14, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        ws.append([]) # Ligne vide

        # Section Résumé
        ws.append(["Indicateur", "Valeur (FC)"])
        summary_rows = [
            ("Total Revenus (Ventes)", data['revenue']),
            ("Coûts Stock (Achats)", data['stock_costs']),
            ("Frais Fonctionnement (Charges)", data['operating_expenses']),
            ("Bénéfice Net", data['net_profit'])
        ]
        
        for row in summary_rows:
            ws.append(row)
        
        # Mise en forme Résumé
        for row in ws.iter_rows(min_row=3, max_row=7, min_col=1, max_col=2):
            for cell in row:
                cell.border = border
                if cell.row == 3:
                    cell.fill = header_fill
                    cell.font = header_font

        # Onglet Détails Ventes
        ws_sales = wb.create_sheet("Détails Ventes")
        ws_sales.append(["ID", "Date", "Mode Paiement", "Montant (FC)"])
        for sale in data['sales']:
            ws_sales.append([sale.id, sale.sale_date.strftime('%d/%m/%Y'), sale.get_payment_method_display(), sale.total])
            
        # Onglet Détails Dépenses
        ws_exp = wb.create_sheet("Détails Charges")
        ws_exp.append(["Date", "Catégorie", "Intitulé", "Montant (FC)"])
        for exp in data['expenses']:
            ws_exp.append([exp.date.strftime('%d/%m/%Y'), exp.category.name, exp.title, exp.amount])

        # Finalisation réponse
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    def to_pdf(data, period_name="Rapport"):
        """Génère un PDF professionnel."""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        
        # Titre
        title_style = ParagraphStyle('TitleStyle', parent=styles['Heading1'], alignment=1, spaceAfter=20)
        elements.append(Paragraph(f"MKARIBU - RAPPORT FINANCIER", title_style))
        elements.append(Paragraph(f"Période : {period_name}", styles['Normal']))
        elements.append(Spacer(1, 12))

        # Tableau Résumé
        table_data = [
            ['INDICATEUR', 'MONTANT (FC)'],
            ['Revenus Totaux', f"{data['revenue']:,.2f}"],
            ['Coûts de Stock', f"{data['stock_costs']:,.2f}"],
            ['Charges de Fonctionnement', f"{data['operating_expenses']:,.2f}"],
            ['BÉNÉFICE NET', f"{data['net_profit']:,.2f}"]
        ]
        
        t = Table(table_data, colWidths=[200, 150])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 4), (-1, 4), 'Helvetica-Bold'),
            ('TEXTCOLOR', (0, 4), (1, 4), colors.red if data['net_profit'] < 0 else colors.green),
        ]))
        elements.append(t)
        
        doc.build(elements)
        buffer.seek(0)
        return buffer
